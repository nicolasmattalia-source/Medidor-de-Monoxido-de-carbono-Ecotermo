import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


st.set_page_config(page_title="Corrección de Ensayos CO", layout="wide")
st.title("Procesador de ensayos de Monóxido de Carbono (CO)")
st.caption(
    "Subí datos del ensayo, planilla de correcciones por medidor y una plantilla de Excel para generar un informe descargable."
)


def read_table(uploaded_file):
    if uploaded_file is None:
        return None

    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file)


def split_cell_ref(cell_ref: str):
    letters = "".join(ch for ch in cell_ref if ch.isalpha()).upper()
    digits = "".join(ch for ch in cell_ref if ch.isdigit())
    if not letters or not digits:
        raise ValueError("Referencia de celda inválida")
    return letters, int(digits)


with st.sidebar:
    st.header("1) Carga de archivos")
    ensayo_file = st.file_uploader(
        "Datos del ensayo (.xlsx o .csv)", type=["xlsx", "xls", "csv"], key="ensayo"
    )
    correccion_file = st.file_uploader(
        "Planilla de correcciones por medidor (.xlsx o .csv)",
        type=["xlsx", "xls", "csv"],
        key="correccion",
    )
    plantilla_file = st.file_uploader(
        "Plantilla Excel de salida (.xlsx)", type=["xlsx"], key="plantilla"
    )

ensayo_df = read_table(ensayo_file)
correccion_df = read_table(correccion_file)

if ensayo_df is None:
    st.info("Esperando archivo de ensayo…")
    st.stop()

st.subheader("2) Configuración de columnas")
col1, col2 = st.columns(2)

with col1:
    st.markdown("#### Ensayo")
    ensayo_cols = list(ensayo_df.columns)
    ensayo_key_col = st.selectbox("Columna identificadora del medidor/modelo", ensayo_cols)
    ensayo_raw_col = st.selectbox("Columna de lectura CO sin corregir", ensayo_cols)

with col2:
    if correccion_df is not None:
        st.markdown("#### Correcciones")
        cor_cols = list(correccion_df.columns)
        cor_key_col = st.selectbox("Columna ID medidor/modelo", cor_cols)
        cor_factor_col = st.selectbox("Columna factor (pendiente)", cor_cols)
        cor_offset_col = st.selectbox("Columna offset (ordenada)", cor_cols)
    else:
        cor_key_col = cor_factor_col = cor_offset_col = None
        st.warning("Cargá la planilla de correcciones para continuar con cálculo automático.")

if correccion_df is None:
    st.stop()

st.subheader("3) Cálculo de datos corregidos")
merged = ensayo_df.merge(
    correccion_df[[cor_key_col, cor_factor_col, cor_offset_col]],
    left_on=ensayo_key_col,
    right_on=cor_key_col,
    how="left",
)

merged[cor_factor_col] = pd.to_numeric(merged[cor_factor_col], errors="coerce").fillna(1.0)
merged[cor_offset_col] = pd.to_numeric(merged[cor_offset_col], errors="coerce").fillna(0.0)
merged[ensayo_raw_col] = pd.to_numeric(merged[ensayo_raw_col], errors="coerce")
merged["co_corregido"] = merged[ensayo_raw_col] * merged[cor_factor_col] + merged[cor_offset_col]

missing_rules = merged[cor_key_col].isna().sum()
if missing_rules:
    st.warning(
        f"{missing_rules} filas no encontraron regla de corrección. Se aplicó factor=1 y offset=0."
    )

st.dataframe(merged.head(50), use_container_width=True)

stats = {
    "promedio": merged["co_corregido"].mean(skipna=True),
    "máximo": merged["co_corregido"].max(skipna=True),
    "mínimo": merged["co_corregido"].min(skipna=True),
}

c1, c2, c3 = st.columns(3)
c1.metric("CO corregido promedio", f"{stats['promedio']:.3f}")
c2.metric("CO corregido máximo", f"{stats['máximo']:.3f}")
c3.metric("CO corregido mínimo", f"{stats['mínimo']:.3f}")

csv_bytes = merged.to_csv(index=False).encode("utf-8")
st.download_button(
    "Descargar datos corregidos (CSV)",
    data=csv_bytes,
    file_name="datos_corregidos.csv",
    mime="text/csv",
)

st.subheader("4) Completar plantilla de Excel")
if plantilla_file is None:
    st.info("Subí una plantilla de Excel para habilitar la exportación final.")
    st.stop()

sheet_name = st.text_input("Hoja destino", value="Sheet1")

with st.expander("Mapeo de celdas de resumen"):
    model_value = st.text_input("Modelo o descripción a escribir", value="Modelo CO")
    model_cell = st.text_input("Celda para modelo", value="B2")
    date_cell = st.text_input("Celda para fecha de proceso", value="B3")
    avg_cell = st.text_input("Celda para promedio CO corregido", value="B4")
    max_cell = st.text_input("Celda para máximo CO corregido", value="B5")

with st.expander("Cargar tabla de resultados en plantilla"):
    start_cell = st.text_input("Celda inicial para tabla", value="A10")
    export_cols = st.multiselect(
        "Columnas a exportar en tabla", options=list(merged.columns), default=[ensayo_key_col, ensayo_raw_col, "co_corregido"]
    )

if st.button("Generar Excel final"):
    try:
        wb = load_workbook(io.BytesIO(plantilla_file.getvalue()))
        if sheet_name not in wb.sheetnames:
            st.error(f"La hoja '{sheet_name}' no existe en la plantilla.")
            st.stop()

        ws = wb[sheet_name]
        ws[model_cell] = model_value
        ws[date_cell] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        ws[avg_cell] = float(stats["promedio"])
        ws[max_cell] = float(stats["máximo"])

        if export_cols:
            start_col_letters, start_row = split_cell_ref(start_cell)
            start_col_index = column_index_from_string(start_col_letters)

            for i, col_name in enumerate(export_cols):
                ws[f"{get_column_letter(start_col_index + i)}{start_row}"] = col_name

            for row_idx, row_data in enumerate(merged[export_cols].itertuples(index=False), start=start_row + 1):
                for col_idx, value in enumerate(row_data, start=start_col_index):
                    ws[f"{get_column_letter(col_idx)}{row_idx}"] = value

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Excel generado correctamente.")
        st.download_button(
            "Descargar Excel final",
            data=output.getvalue(),
            file_name="reporte_co_corregido.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        st.error(f"No se pudo generar el Excel: {exc}")
